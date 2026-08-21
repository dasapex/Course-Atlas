#!/usr/bin/env python3
"""Extract the FEU Tech BSCE curriculum workbook into Hugo data JSON.

The workbook is arranged as four year blocks, with five columns per year:
course code, course name, units, co-requisites, and prerequisites. The script
also combines explicitly paired lecture/laboratory subjects into one card.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

import openpyxl


SCRIPT_DIR = Path(__file__).resolve().parent

# When installed at scripts/extract_curriculum.py, the project root is its
# parent. When run as a standalone downloaded file, use its own directory.
PROJECT_ROOT = SCRIPT_DIR.parent if SCRIPT_DIR.name == "scripts" else SCRIPT_DIR
DEFAULT_SOURCE = PROJECT_ROOT / "source-data" / "BSCE-FEUIT-CurriculumTree.xlsx"
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "courses.json"

# (column offset, curriculum term, first data row, last data row)
BLOCKS = [
    (0, 1, 3, 12),
    (0, 2, 14, 22),
    (0, 3, 24, 33),
    (5, 4, 3, 12),
    (5, 5, 14, 22),
    (5, 6, 24, 33),
    (10, 7, 3, 12),
    (10, 8, 14, 22),
    (10, 9, 24, 33),
    (15, 10, 3, 12),
    (15, 11, 14, 22),
    (15, 12, 24, 33),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert the FEU Tech BSCE curriculum workbook to Hugo JSON."
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=DEFAULT_SOURCE,
        help=f"Input workbook (default: {DEFAULT_SOURCE})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output JSON file (default: {DEFAULT_OUTPUT})",
    )
    return parser.parse_args()


def split_codes(value: Any) -> list[str]:
    """Turn a comma-separated prerequisite cell into normalized IDs."""
    if not value:
        return []
    return [part.strip().lower() for part in str(value).split(",") if part.strip()]


def clean_number(value: Any) -> int | float:
    """Keep whole unit values as integers for cleaner JSON and labels."""
    number = float(value or 0)
    return int(number) if number.is_integer() else number


def category_for(code: str) -> str:
    if code.startswith("CE"):
        return "ce"
    if code.startswith("COE"):
        return "coe"
    # GED and NSTP subjects share the general-education display group.
    return "ged"


def read_raw_courses(source: Path) -> list[dict[str, Any]]:
    """Read all 12 curriculum terms from the arranged worksheet."""
    if not source.is_file():
        raise FileNotFoundError(
            f"Workbook not found: {source}\n"
            "Place it in source-data/ or provide --source PATH."
        )

    worksheet = openpyxl.load_workbook(source, data_only=True).active
    courses: list[dict[str, Any]] = []

    for base, term, first_row, last_row in BLOCKS:
        for row in range(first_row, last_row + 1):
            code_value = worksheet.cell(row, base + 1).value
            if not code_value:
                continue

            code = str(code_value).strip()
            course_id = code.lower()
            name_value = worksheet.cell(row, base + 2).value
            units_value = worksheet.cell(row, base + 3).value

            corequisites = [
                item
                for item in split_codes(worksheet.cell(row, base + 4).value)
                if item != course_id
            ]
            prerequisites = [
                item
                for item in split_codes(worksheet.cell(row, base + 5).value)
                if item != course_id and item not in corequisites
            ]

            courses.append(
                {
                    "id": course_id,
                    "code": code,
                    "name": " ".join(str(name_value or "").split()),
                    "units": clean_number(units_value),
                    "cat": category_for(code),
                    "term": term,
                    "prereq": prerequisites,
                    "coreq": corequisites,
                }
            )

    if not courses:
        raise ValueError(
            "No courses were found. Confirm that the workbook uses the "
            "'By Year (Arranged)' layout supplied with this project."
        )

    return courses


def merge_lecture_labs(courses: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Combine only labs explicitly linked as a lecture's co-requisite."""
    by_id = {course["id"]: course for course in courses}
    lab_to_lecture: dict[str, str] = {}

    for course in courses:
        for corequisite_id in course["coreq"]:
            lab = by_id.get(corequisite_id)
            if (
                lab
                and corequisite_id.endswith("l")
                and "LAB" in lab["name"].upper()
            ):
                lab_to_lecture[corequisite_id] = course["id"]

    merged: list[dict[str, Any]] = []

    for course in courses:
        if course["id"] in lab_to_lecture:
            continue

        lab_ids = [
            lab_id
            for lab_id, lecture_id in lab_to_lecture.items()
            if lecture_id == course["id"]
        ]
        labs = [by_id[lab_id] for lab_id in lab_ids]
        item = dict(course)

        item["lectureUnits"] = clean_number(course["units"])
        item["labUnits"] = clean_number(sum(lab["units"] for lab in labs))
        item["units"] = clean_number(course["units"] + item["labUnits"])

        if labs:
            item["name"] = (
                item["name"].replace(" (LEC)", "").replace(" (LECTURE)", "")
            )
            item["prereq"] = list(
                dict.fromkeys(
                    course["prereq"]
                    + [reference for lab in labs for reference in lab["prereq"]]
                )
            )

        item["coreq"] = [
            reference for reference in course["coreq"] if reference not in lab_ids
        ]
        merged.append(item)

    # Redirect dependencies that formerly pointed to a merged lab card.
    for course in merged:
        course["prereq"] = list(
            dict.fromkeys(
                lab_to_lecture.get(reference, reference)
                for reference in course["prereq"]
                if lab_to_lecture.get(reference, reference) != course["id"]
            )
        )
        course["coreq"] = list(
            dict.fromkeys(
                lab_to_lecture.get(reference, reference)
                for reference in course["coreq"]
                if lab_to_lecture.get(reference, reference) != course["id"]
            )
        )

    return merged


def validate_courses(courses: list[dict[str, Any]]) -> None:
    """Fail early instead of allowing Hugo to build an incomplete map."""
    ids = [course["id"] for course in courses]
    duplicate_ids = sorted({course_id for course_id in ids if ids.count(course_id) > 1})
    known_ids = set(ids)
    missing_references = sorted(
        {
            reference
            for course in courses
            for reference in course["prereq"] + course["coreq"]
            if reference not in known_ids
        }
    )

    if duplicate_ids:
        raise ValueError(f"Duplicate course IDs: {', '.join(duplicate_ids)}")
    if missing_references:
        raise ValueError(
            "Unknown prerequisite/co-requisite IDs: "
            + ", ".join(missing_references)
        )
    if any(not course["name"] for course in courses):
        raise ValueError("One or more extracted courses has no subject name.")


def write_hugo_data(courses: list[dict[str, Any]], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(courses, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    args = parse_args()
    source = args.source.resolve()
    output = args.output.resolve()

    raw_courses = read_raw_courses(source)
    courses = merge_lecture_labs(raw_courses)
    validate_courses(courses)
    write_hugo_data(courses, output)

    total_units = sum(course["units"] for course in courses)
    merged_pairs = sum(1 for course in courses if course["labUnits"])

    print(f"Source: {source}")
    print(f"Output: {output}")
    print(f"Extracted source subjects: {len(raw_courses)}")
    print(f"Generated course cards: {len(courses)}")
    print(f"Merged lecture/lab pairs: {merged_pairs}")
    print(f"Total curriculum units: {total_units:g}")


if __name__ == "__main__":
    main()